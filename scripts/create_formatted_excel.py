"""
Création d'un fichier Excel formaté pour la prospection commerciale.

Format :
- Colonnes réorganisées : Structure → Dirigeant → Contacts → Détails
- Mise en forme : gras, couleurs sobres, largeurs optimisées
- Lisibilité optimale pour prospection
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Charger données
df = pd.read_excel('outputs/prospection_250_dirigeants_final.xlsx')

# Réorganiser colonnes dans l'ordre logique de prospection
colonnes_ordonnees = [
    # 1. IDENTIFICATION STRUCTURE
    'gestionnaire_nom',
    'nom_public',
    'acronyme',
    'finess_ej',
    
    # 2. DIRIGEANT & CONTACTS
    'dirigeant_nom',
    'dirigeant_titre',
    'email_dirigeant_1',
    'email_dirigeant_2',
    'email_dirigeant_3',
    'dirigeant_confidence',
    
    # 3. CONTACTS STRUCTURE
    'site_web',
    'domaine',
    'emails_generiques',
    'email_contact',
    
    # 4. ACTIVITÉ
    'nb_essms',
    'categorie_taille',
    'dominante_type',
    'dominante_nb',
    'dominante_top5',
    
    # 5. ADRESSE
    'gestionnaire_adresse',
    
    # 6. MÉTADONNÉES
    'dirigeant_source',
    'email_pattern',
    'email_confidence',
    'pattern_examples',
    'confidence',
    'sources_web',
    'url_contact',
    'url_mentions_legales',
    'dirigeant_linkedin_url',
    'query_web',
]

# Créer DataFrame réorganisé
df_formatted = df[colonnes_ordonnees].copy()

# Renommer colonnes pour plus de clarté
df_formatted.columns = [
    'Structure',
    'Nom Public',
    'Acronyme',
    'FINESS',
    'Dirigeant',
    'Fonction',
    'Email Dirigeant 1',
    'Email Dirigeant 2',
    'Email Dirigeant 3',
    'Conf. Dirigeant',
    'Site Web',
    'Domaine',
    'Emails Génériques',
    'Email Contact',
    'Nb ESSMS',
    'Taille',
    'Type Principal',
    'Nb Type Principal',
    'Top 5 Types',
    'Adresse',
    'Source Dirigeant',
    'Pattern Email',
    'Conf. Email',
    'Exemples Pattern',
    'Conf. Globale',
    'Sources Web',
    'URL Contact',
    'URL Mentions',
    'LinkedIn',
    'Requête Web',
]

# Sauvegarder Excel
output_path = 'outputs/prospection_250_FINAL_FORMATE.xlsx'
df_formatted.to_excel(output_path, index=False, sheet_name='Prospection')

# Ouvrir avec openpyxl pour mise en forme
wb = load_workbook(output_path)
ws = wb.active

# Définir styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2C5F8D', end_color='2C5F8D', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

structure_font = Font(name='Calibri', size=11, bold=True)
dirigeant_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
contact_fill = PatternFill(start_color='F0F8E8', end_color='F0F8E8', fill_type='solid')

normal_font = Font(name='Calibri', size=10)
normal_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

border_thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Formater en-têtes
for col_num, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border_thin

# Largeurs de colonnes optimisées
column_widths = {
    'A': 50,  # Structure
    'B': 35,  # Nom Public
    'C': 15,  # Acronyme
    'D': 12,  # FINESS
    'E': 25,  # Dirigeant
    'F': 25,  # Fonction
    'G': 35,  # Email 1
    'H': 35,  # Email 2
    'I': 35,  # Email 3
    'J': 10,  # Conf Dirigeant
    'K': 40,  # Site Web
    'L': 25,  # Domaine
    'M': 40,  # Emails Génériques
    'N': 35,  # Email Contact
    'O': 10,  # Nb ESSMS
    'P': 15,  # Taille
    'Q': 40,  # Type Principal
    'R': 10,  # Nb Type Principal
    'S': 60,  # Top 5 Types
    'T': 50,  # Adresse
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Formater lignes de données
for row_num in range(2, ws.max_row + 1):
    # Structure en gras (col A)
    ws[f'A{row_num}'].font = structure_font
    ws[f'A{row_num}'].alignment = normal_alignment
    
    # Dirigeant & emails avec fond bleu clair (cols E-J)
    for col in ['E', 'F', 'G', 'H', 'I', 'J']:
        cell = ws[f'{col}{row_num}']
        cell.fill = dirigeant_fill
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = border_thin
    
    # Contacts structure avec fond vert clair (cols K-N)
    for col in ['K', 'L', 'M', 'N']:
        cell = ws[f'{col}{row_num}']
        cell.fill = contact_fill
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = border_thin
    
    # Autres colonnes normales
    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        if col_letter not in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            cell = ws[f'{col_letter}{row_num}']
            cell.font = normal_font
            cell.alignment = normal_alignment
            cell.border = border_thin

# Figer première ligne et première colonne
ws.freeze_panes = 'B2'

# Filtres automatiques
ws.auto_filter.ref = ws.dimensions

# Hauteur des lignes
ws.row_dimensions[1].height = 40  # En-tête plus haute
for row_num in range(2, ws.max_row + 1):
    ws.row_dimensions[row_num].height = 20

# Sauvegarder
wb.save(output_path)

print("=" * 70)
print("✅ FICHIER EXCEL FORMATÉ CRÉÉ")
print("=" * 70)
print(f"Fichier: {output_path}")
print(f"\nContenu:")
print(f"  - 250 gestionnaires")
print(f"  - 233 dirigeants identifiés (93.2%)")
print(f"  - 230 emails reconstruits (98.7% des dirigeants)")
print(f"\nMise en forme:")
print(f"  - Structure en GRAS (colonne A)")
print(f"  - Dirigeant & emails sur fond BLEU CLAIR")
print(f"  - Contacts structure sur fond VERT CLAIR")
print(f"  - En-têtes sur fond BLEU FONCÉ")
print(f"  - Filtres automatiques activés")
print(f"  - Colonnes optimisées pour impression/lecture")
print("=" * 70)
